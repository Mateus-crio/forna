import openpyxl
openpyxl.Workbook('Dados.xlsx') 
book = openpyxl.load_workbook('Dados.xlsx') 
print(book.sheetnames) 


book.create_sheet("Custo de Produção")
print(book.sheetnames)

book["Custo de Produção"]
frutas_page = book["Custo de Produção"]

frutas_page.append(["Valor", "Tempo"])
frutas_page.append(["R$ 5.566.899", "10"])
frutas_page.append(["R$ 139.543.156", "15"])
frutas_page.append(["R$ 154.392.755", "50"])

book.save("Dados.xlsx")