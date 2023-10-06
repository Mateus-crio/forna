import openpyxl
openpyxl.load_workbook("Dados.xlsx")
book = openpyxl.load_workbook("Dados.xlsx")

print(book.sheetnames) 
book.create_sheet("Valores da matéria-prima")

frutas_page = book["Valores da matéria-prima"]

frutas_page.append(["Nome da Matéria-prima", "Valor","Local"])
frutas_page.append(["Aço", "R$ 100.369	","São Paulo"])
frutas_page.append(["Titanio", "R$ 543.156","Santa Catarina"])
frutas_page.append(["Metal", "R$ 392.755","Rio de Janeiro"])

book.save("Dados.xlsx")