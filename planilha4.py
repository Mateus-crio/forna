import openpyxl
openpyxl.load_workbook("Dados.xlsx")
book = openpyxl.load_workbook("Dados.xlsx")

print(book.sheetnames) 
book.create_sheet("Valores de Fretes")

frutas_page = book["Valores de Fretes"]

frutas_page.append(["Valor","Local"])
frutas_page.append(["R$ 33.369", "São Paulo"])
frutas_page.append(["R$ 21.156", "Santa Catarina"])
frutas_page.append(["R$ 43.755", "Rio de Janeiro"])

book.save("Dados.xlsx")